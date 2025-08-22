from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
from PIL import Image
import io
import streamlit as st
import pandas as pd
import base64

# =========================
# KONFIGURASI DASAR
# =========================
st.set_page_config(layout="wide", page_title="Dashboard RAB Gardu", page_icon="Bagian_Gardu/Logo_PLN.png")

# =========================
# ========== LOGIN =========
# =========================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

st.markdown(
    """
    <style>
    .stForm > div {
        border: 2px solid #00468C !important; /* Biru PLN */
        border-radius: 8px;
        max-width: 550px;
        background-color: #00000; /* Contoh: kuning solid */
        padding: 20px;
        margin-left: auto;
        margin-right: auto;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Data akun (bisa kamu ganti nanti)
VALID_USERS = {
    "admin": "admin123",
    "plnuser": "pln2024"
}
def get_base64_of_bin_file(bin_file):
    with open(bin_file, 'rb') as f:
        data = f.read()
    return base64.b64encode(data).decode()
    
bin_str = get_base64_of_bin_file("Bagian_Gardu/Penyulang.jpeg")
# Tulis CSS di dalam string triple quotes, f-string

def login():
    # Tambahkan CSS background + overlay

    st.markdown(
        """
        <style>
        .stApp h1 {
            color: #ffffff; /* Putih */
            text-shadow: 1px 1px 4px rgba(0,0,0,0.6); /* Tambahkan bayangan supaya lebih terbaca */
        }
        </style>
        """,
        unsafe_allow_html=True
    )
    st.markdown(
        """
        <style>
        /* Ubah warna kotak form */
        .stForm > div {
            background-color: #FFFFFF !important; /* Kuning PLN */
            border: 2px solid #FFD700; /* Kalau mau border kuning juga */
            border-radius: 8px;
            padding: 20px;
        }
        </style>
        """,
        unsafe_allow_html=True
    )
    st.markdown(
        """
        <h1 style="text-align: center; color: white; text-shadow: 1px 1px 4px rgba(0,0,0,0.6);">
            Login Sistem RAB & Visualisasi Gardu
        </h1>
        """,
        unsafe_allow_html=True
    )
    
    with st.form("login_form"):
        st.markdown(
            """
            <style>
            /* Border input text */
            input[type="text"], input[type="password"] {
                border: 2px solid #000000 !important;  /* Hitam PLN */
                border-radius: 5px;
            }
        
            /* Border input text saat fokus */
            input[type="text"]:focus, input[type="password"]:focus {
                border: 2px solid #FFCC00 !important; /* Warna kuning lebih terang saat fokus */
                outline: none;
                box-shadow: 0 0 5px rgba(255,204,0,0.8);
            }
            </style>
            """,
            unsafe_allow_html=True
        )
        username = st.text_input("Username").strip()
        password = st.text_input("Password", type="password").strip()
        submitted = st.form_submit_button("Login")

        if submitted:
            if username in VALID_USERS and VALID_USERS[username] == password:
                st.session_state.logged_in = True
                st.success("Login berhasil. Silakan lanjut.")
                st.rerun()
            else:
                st.error("Username atau password salah")



if not st.session_state.logged_in:
    login()
    st.markdown(
        f"""
        <style>
        .stApp {{
            background-image: url("data:image/jpeg;base64,{bin_str}");
            background-size: cover;
            background-position: center;
        }}
        .stApp::before {{
            content: "";
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0, 70, 140, 0.5);
            z-index: 0;
        }}
        .main > div {{
            position: relative;
            z-index: 1;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )
    
    st.stop()
    
with st.sidebar:
    st.image("Bagian_Gardu/Logo_PLN.png", width=120)
    if st.button("Logout"):
        st.session_state.logged_in = False
        st.rerun()
        # Logo BUMN di pojok kiri bawah
        st.markdown(
            """
            <style>
            .bumn-logo-container {
                position: absolute;
                bottom: 10px;   /* Jarak dari bawah sidebar */
                left: 10px;     /* Jarak dari kiri sidebar */
            }
            .bumn-logo-container img {
                width: 60px;    /* Ukuran logo BUMN */
            }
            </style>
            <div class="bumn-logo-container">
                <img src="Bagian_Gardu/BUMN.png">
            </div>
            """,
            unsafe_allow_html=True
        )

# =========================
# DATA HARGA & GAMBAR
# =========================
st.markdown(
    """
    <style>
    .stApp h1 {
        color: #00000; 
        text-shadow: 1px 1px 4px rgba(0,0,0,0.6); /* Tambahkan bayangan supaya lebih terbaca */
    }
    </style>
    """,
    unsafe_allow_html=True
)
st.title("Dahsboard Sistem RAB Beserta Visualisasi")
gardu_data = {
    "Gardu Cantol": {
        "material": {
            "Pasang TM - 1 B/H ( TIANG TUMPU )": [105620, "Set"],
            "PENGADAAN DAN PEMASANGAN PENGHALANG PANJAT": 72126,
            "PENGADAAN DAN PEMASANGAN TANDA KILAT KECIL": 59845,
            "Penomoran Tiang TM": 13888,
            "Pasang rangka gardu Cantol": 350124,
            "Pemasangan hardware, isolator, pipa, PGTM/TR Cantol": 446693,
            "Pasang trafo 100 kVA":  626657,
            "Pasang Arrester + Jumper": 46824,
            "Pasang Cut Out + Jumper": 46824,
            "Pasang Cover Arrester": 10498,
            "Pasang Cover Cut Out": 10498,
            "Pasang Cover Bushing Trafo": 19498,
            "Pembuatan pondasi gardu Cantol": 1894660,
            "Pembuatan Nama dan logo gardu": 48887,
            "Pasang Pentanahan TM Gardu Cantol": 228572,
            "Pasang Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat": 163122,
            "Pasang Kbl Single Core 20 kV lengkap Pipa Pelindung": 320891,
            "Tagging GPS": 3437,
            "Dokumentasi / Foto 1 album": 78539,
            "Biaya Ongkos Angkut material Pick Up": 268257,
            "Biaya Gambar Pelaksanaan Pada Millar / Kalkir 60 mikron A1 (Blue Print 4 lbr)": 288232,
            "TIANG BETON BULAT 12m/350 daN (terpasang) + Lansir": 5572770,
            "CONDUCTOR;AAAC-S;70mm2": 15610,
            "CABLE PWR;N2XSY 1x35mm2;20kV;UG": 132390,
            "CABLE PWR;NYY;1X70mm2;0,6/1kV;OH": 117380,
            "CABLE PWR;NYY;1X120mm2;0,6/1kV;OH": 428735,
            "CABLE PWR ACC;CABLE SHOE CU ID 1H 70mm2": 69000,
            "CABLE PWR ACC;CABLE SHOE CU ID 1H 240mm2": 238000,
            "CABLE PWR ACC;CABLE SHOE AL ID 1H 70mm2": 62600,
            "CUT 20kV;24kV;R-VO;7-B;70-150mm2;PRS": 34500,
            "TRA DIS;24kV;400/230;25kVA;DYNS;00": 82638,
            "LVSB;DIST;3P;400V;250A;2Line;OH": 14408190,
            "Line Post;24kV;12.5kN;Polymer SIR": 219040,
            "Insulated Top Ties": 69500,
            "LA;24-24kV;1kA;Polymer": 606990,
            "CUT OUT;24kV;100A;Polymer-125kV": 1164370.00
        },
         "image": "Bagian_Gardu/GarduCantolRevisi/StrukturGarduCantol.jpg"
    },
    "Gardu Tembok": {
        "material": {
            "PEMASANGAN CUBICLE PROTEKSI TRAFO (PB)":  313062,
            "PASANG GROUND FAULT DETECTOR (GFD) DI GARDU TEMBOK":   51443,
            "PASANG PHBTR INDOOR 4 JURUSAN / 6 JURUSAN":   280161,
            "PEMASANGAN KABEL NYY DARI TRAFO KE PHBTR DI GARDU TEMBOK":   3286,
            "PASANG SINGLE-CORE, INDORESIN & ELASTIMOLD":   531965,
            "PEMASANGAN INSTALASI PENTANAHAN DI GARDU TEMBOK":  53500,
            "PLAT LANTAI KABEL LKP PENUTUP LUBANG DENGAN AQUAPROOF":  174649,
            "PASANG TRAFO 400 kVA GARDU TEMBOK":  600000,
            "CONDUCTOR;CU;50MM2;19.28KN":  54084,
            "CONN;;PG;CU;50MM2;BOLT":  29679,
            "PEMBUATAN GAMBAR PELAKSANAAN PADA KERTAS HVS 80 MGs UKURAN A3 ( 4 LEMBAR )":  150000,
            "FOTO DOKUMENTASI":  150000,
            "CABLE PWR;N2XSY;1X35mm2;20kV;UG":  132300,
            "CABLE PWR;NYY;1X240mm2;0,6/1kV;OH":  420735,
            "CABLE PWR ACC;CABLE SHOE CU ID 1H 240mm2":  238000,
            "TRF DIS;D3;20KV/400V;3P;400KVA;DYN5;ID":  118520636,
            "CUB;N ISO;TP;24KV;630A;16KA":  52455940,
            "LVSB;DIST;3P;380V;1000A;6LINE;ID":  49252470,
            "INDIKATOR GANGGUAN KABEL BAWAH TANAH (GFD)":  1560000,
            "FUSE;20/24kV;16A;TUBE;D24mm":  1564779
        },
        "image": "/content/drive/MyDrive/Bagian_Gardu/GarduTembok/StukturGarduTembok.JPG"
    },
    "Gardu Portal": {
        "material": {
            "Pasang TM - 1 B/H ( TIANG TUMPU )": 105620,
            "PENGADAAN DAN PEMASANGAN PENGHALANG PANJAT": 72126,
            "PENGADAAN DAN PEMASANGAN TANDA KILAT KECIL": 59845,
            "Penomoran Tiang TM": 13888,
            "Pasang rangka gardu Portal": 525186,
            "Pemasangan hardware, isolator, pipa, PGTM/TR Portal": 1026850,
            "Pasang trafo 200 - 250 kVA": 7953990,
            "Pasang Arrester + Jumper": 46824,
            "Pasang Cut Out + Jumper": 46824,
            "Pasang Cover Arrester": 10498,
            "Pasang Cover Cut Out": 10498,
            "Pasang Cover Bushing Trafo": 19498,
            "Pembuatan pondasi gardu Portal": 3948992,
            "Pembuatan Nama dan logo gardu": 48887,
            "Pasang Pentanahan Ganda TM Gardu Portal": 365715,
            "Pasang Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat": 163122,
            "Tagging GPS": 3437,
            "Dokumentasi / Foto 1 album": 78539,
            "Biaya Ongkos Angkut material Pick Up": 268257,
            "Biaya Gambar Pelaksanaan Pada Millar / Kalkir 60 mikron A1 (Blue Print 4 lbr)": 288232,
            "TIANG BETON BULAT 12m/350 daN (terpasang) + Lansir": 5572770,

            # Kabel
            "CONDUCTOR;AAAC-S;150mm2": 27240,
            "CONDUCTOR;AAAC-S;70mm2": 15610,
            "CABLE PWR;N2XSY 1x70mm2;20kV;UG": 117380,
            "CABLE PWR;NYY;1X240mm2;0,6/1kV;OH": 428735,

            # Konektor
            "CABLE PWR ACC;CABLE SHOE CU ID 1H 70mm2": 69000,
            "CABLE PWR ACC;CABLE SHOE CU ID 1H 240mm2": 238000,
            "CABLE PWR ACC;CABLE SHOE AL ID 1H 70mm2": 62600,
            "CONN;20kV;24kV;H-35-70/70-150mm2;PRS": 34500,

            # Perlengkapan gardu
            "TRF DIS;24kV;400/230;25kVA;DYNS;00": 82638,
            "LVSB;DIST;3P;400V;630A;4Line;OD": 28466360,

            # Perlengkapan jaringan
            "Line Post;24kV;12.5kN;Polymer SIR": 219040,
            "INSULATED TOP TIES": 69500,

            # Pelindung/pengontrol
            "LA;24-24kV;1kA;Polymer": 606990,
            "CUT OUT;24kV;6-100A;POLYMER-125kV": 1164370,
            "CUT OUT ACC;FUSE LINK 26 kV 8A": 19980
        },
        "image": "Bagian_Gardu/GarduPortalRevisi/GARDU PORTAL-Full lengkap GP.drawio.png"
    }
}

gambat_material_per_gardu = {
    "Gardu Tembok": {
        "PEMASANGAN CUBICLE PROTEKSI TRAFO (PB)":  "Bagian_Gardu/GarduTembok/GARDU TEMBOK-Cubicle GT.drawio.png",
        "PASANG PHBTR INDOOR 4 JURUSAN / 6 JURUSAN":   "Bagian_Gardu/GarduTembok/GARDU TEMBOK-PHBTR GT.drawio.png",
        "PEMASANGAN KABEL NYY DARI TRAFO KE PHBTR DI GARDU TEMBOK":   "Bagian_Gardu/GarduTembok/GARDU TEMBOK-Kabel Trafo PHBTR GT.drawio.png",
        "PLAT LANTAI KABEL LKP PENUTUP LUBANG DENGAN AQUAPROOF":  "Bagian_Gardu/GarduTembok/GARDU TEMBOK-Lantai GT.drawio.png",
        "PEMASANGAN INSTALASI PENTANAHAN DI GARDU TEMBOK":  "Bagian_Gardu/GarduTembok/GARDU TEMBOK-Pentanahan GT.drawio.png",
        "PASANG TRAFO 400 kVA GARDU TEMBOK":  "Bagian_Gardu/GarduTembok/GARDU TEMBOK-Trafo GT.drawio.png",
        "LVSB;DIST;3P;380V;1000A;6LINE;ID":  "Bagian_Gardu/GarduTembok/GARDU TEMBOK-LVSB GT.drawio.png"
    },
    "Gardu Cantol": {
        "Pasang TM - 1 B/H ( TIANG TUMPU )": "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-TM - 1 B_H ( TIANG TUMPU ) GC.png",
        "PENGADAAN DAN PEMASANGAN PENGHALANG PANJAT": "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-PENGHALANG PANJAT GC.png",
        "PENGADAAN DAN PEMASANGAN TANDA KILAT KECIL": "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-TANDA KILAT KECIL GC.png",
        "Penomoran Tiang TM": "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Penomoran Tiang TM GC.png",
        "Pasang rangka gardu Cantol": "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Rangka gardu cantol GC.png",
        "Pemasangan hardware, isolator, pipa, PGTM/TR Cantol": "Bagian_Gardu/GarduCantolRevisi/Revisi_GARDU CANTOL-hardware, isolator, pipa, PGTM_TR Cantol GC.drawio.png",
        "Pasang trafo 100 kVA": "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-trafo 200 - 250 kVA GC(1).png",
        "LVSB;DIST;3P;400V;250A;2Line;OH": "Bagian_Gardu/GarduCantolRevisi/LVSB;DIST;3P;400V;250A;2Line;OH.png",
        "Pasang Arrester + Jumper": [
            "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Arrester + Jumper 1 GC.png",
            "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Arrester + Jumper 2 GC.png",
            "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Arrester + Jumper 3 GC.png"
        ],
        "Pasang Cut Out + Jumper": [
            "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL- Cut Out + Jumper 1 GC.png",
            "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL- Cut Out + Jumper 2 GC.png",
            "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL- Cut Out + Jumper 3 GC.png"
        ],
        "Pasang Cover Bushing Trafo": [
            "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Cover Bushing Trafo 1 GC.png",
            "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Cover Bushing Trafo 2 GC.png",
            "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Cover Bushing Trafo 3 GC.png"
        ],
        "Pembuatan pondasi gardu Cantol": "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Pondasi gardu cantol GC.png",
        "Pembuatan Nama dan logo gardu": "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Nama dan logo GC.png",
        "Pasang Pentanahan TM Gardu Cantol": "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Pentanahan TM Gardu Cantol GC.png",
        "Pasang Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat": [
            "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat 1 GC.png",
            "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat 2 GC.png",
        ],
        "TIANG BETON BULAT 12m/350 daN (terpasang) + Lansir": "Bagian_Gardu/GarduCantolRevisi/Copy of GARDU CANTOL-Tiang GC.png"
    },
    "Gardu Portal":{
        "Pasang TM - 1 B/H ( TIANG TUMPU )": [
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-TM - 1 B_H ( TIANG TUMPU ) 1 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-TM - 1 B_H ( TIANG TUMPU ) 2 GP.drawio.png"
        ],
        "PENGADAAN DAN PEMASANGAN PENGHALANG PANJAT": [
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Penghalang Panjat 1 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Penghalang Panjat 2 GP.drawio.png"
        ],
        "PENGADAAN DAN PEMASANGAN TANDA KILAT KECIL": [
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Tanda Kilat Kecil 1 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Tanda Kilat Kecil 2 GP.drawio.png"
        ],
        "Penomoran Tiang TM": [
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Penomoran Tiang 1 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Penomoran Tiang 2 GP.drawio.png"
        ],
        "Pasang rangka gardu Portal": "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Rangka Gardu Portal GP.drawio.png",
        "Pemasangan hardware, isolator, pipa, PGTM/TR Portal": "Bagian_Gardu/GarduPortalRevisi/Revisi_Pemasangan hardware, isolator, pipa, PGTM_TR Portal.png",
        "Pasang trafo 200 - 250 kVA": "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Trafo 200 - 250 kVA GP.drawio.png",
        "Pasang Arrester + Jumper": [
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Arrester + Jumper 1 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Arrester + Jumper 2 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Arrester + Jumper 3 GP.drawio.png"
        ],
        "Pasang Cut Out + Jumper": [
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Cut Out + Jumper 1 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Cut Out + Jumper 2 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Cut Out + Jumper 3 GP.drawio.png"
        ],
        "Pasang Cover Bushing Trafo": [
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Cover Bushing Trafo 1 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Cover Bushing Trafo 2 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Cover Bushing Trafo 3 GP.drawio.png"
        ],
        "Pembuatan pondasi gardu Portal": "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-pondasi gardu Portal GP.drawio.png",
        "Pembuatan Nama dan logo gardu": "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Nama dan Logo Gardu GP.drawio.png",
        "Pasang Pentanahan Ganda TM Gardu Portal": "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Pentanahan Ganda TM Gardu Portal GP.drawio.png",
        "Pasang Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat": [
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat 1 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat 2 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Revisi_Jaringan TR 4 GP.png",
            "Bagian_Gardu/GarduPortalRevisi/Revisi_Jaringan TR 3 GP.png"
        ],
        "LVSB;DIST;3P;400V;630A;4Line;OD": "Bagian_Gardu/GarduPortalRevisi/LVSB;DIST;3P;400V;630A;4Line;OD.png",
        "TIANG BETON BULAT 12m/350 daN (terpasang) + Lansir": [
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-TIANG BETON BULAT 12m_350 daN (terpasang) + Lansir 1 GP.drawio.png",
            "Bagian_Gardu/GarduPortalRevisi/Copy of GARDU PORTAL-TIANG BETON BULAT 12m_350 daN (terpasang) + Lansir 2 GP.drawio.png"
        ]
    }
}
# Jumlah maksimal per material untuk Gardu Portal
pilihan_jumlah_per_material = {
    "Gardu Tembok": {
        "material": {
            "Isolator Tumpu 20 kV": 50000,
            "Lightning Arrester": 75000,
            "Trafo 100 kVA": 5000000,
        }
    },
    "Gardu Cantol": {
        "material": {
            "Pasang TM - 1 B/H ( TIANG TUMPU )": [0,1],
            "PENGADAAN DAN PEMASANGAN PENGHALANG PANJAT": [0,1],
            "PENGADAAN DAN PEMASANGAN TANDA KILAT KECIL": [0,1],
            "Penomoran Tiang TM": [0,1],
            "Pasang rangka gardu Cantol": [0,1],
            "Pemasangan hardware, isolator, pipa, PGTM/TR Cantol": [0,1],
            "Pasang trafo 100 kVA": [0,1],
            "Pasang Arrester + Jumper": [0,1,2,3],
            "Pasang Cut Out + Jumper": [0,1,2,3],
            "Pasang Cover Arrester": [0,1,2,3],
            "Pasang Cover Cut Out": [0,1,2,3],
            "Pasang Cover Bushing Trafo": [0,1,2,3],
            "Pembuatan pondasi gardu Cantol": [0,1],
            "Pembuatan Nama dan logo gardu": [0,1],
            "Pasang Pentanahan TM Gardu Cantol": [0,1],
            "Pasang Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat": [0,1],
            "Pasang Kbl Single Core 20 kV lengkap Pipa Pelindung": [0,1,2],
            "Tagging GPS": [0,1],
            "Dokumentasi / Foto 1 album": [0,1],
            "Biaya Ongkos Angkut material Pick Up": [0,1],
            "Biaya Gambar Pelaksanaan Pada Millar / Kalkir 60 mikron A1 (Blue Print 4 lbr)": [0,1],
            "TIANG BETON BULAT 12m/350 daN (terpasang) + Lansir": [0,1],
            "CUT 20kV;24kV;R-VO;7-B;70-150mm2;PRS": [0,1],
            "TRA DIS;24kV;400/230;25kVA;DYNS;00": [0,1],
            "LVSB;DIST;3P;400V;250A;2Line;OH": [0,1],
            "Line Post;24kV;12.5kN;Polymer SIR": [0,1,2,3],
            "Insulated Top Ties": [0,1,2,3],
            "LA;24-24kV;1kA;Polymer": [0,1,2,3],
            "CUT OUT;24kV;100A;Polymer-125kV": [0,1,2,3]
        }
    },
    "Gardu Portal": {
        "material": {
            "Pasang TM - 1 B/H ( TIANG TUMPU )": [0,1,2],
            "PENGADAAN DAN PEMASANGAN PENGHALANG PANJAT": [0,1,2],
            "PENGADAAN DAN PEMASANGAN TANDA KILAT KECIL": [0,1,2],
            "Penomoran Tiang TM": [0,1,2],
            "Pasang rangka gardu Portal": [0,1],
            "Pemasangan hardware, isolator, pipa, PGTM/TR Portal": [0,1],
            "Pasang trafo 200 - 250 kVA": [0,1],
            "Pasang Arrester + Jumper": [0,1,2,3],
            "Pasang Cut Out + Jumper": [0,1,2,3],
            "Pasang Cover Arrester": [0,1,2,3],
            "Pasang Cover Cut Out": [0,1,2,3],
            "Pasang Cover Bushing Trafo": [0,1,2,3],
            "Pembuatan pondasi gardu Portal": [0,1],
            "Pembuatan Nama dan logo gardu": [0,1],
            "Pasang Pentanahan Ganda TM Gardu Portal": [0,1],
            "Pasang Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat": [0,1,2,3,4],
            "Tagging GPS": [0,1],
            "Dokumentasi / Foto 1 album": [0,1],
            "Biaya Ongkos Angkut material Pick Up": [0,1],
            "Biaya Gambar Pelaksanaan Pada Millar / Kalkir 60 mikron A1 (Blue Print 4 lbr)": [0,1],
            "TIANG BETON BULAT 12m/350 daN (terpasang) + Lansir": [0,1,2],

            # Perlengkapan gardu
            "TRF DIS;24kV;400/230;25kVA;DYNS;00": [0,1],
            "LVSB;DIST;3P;400V;630A;4Line;OD": [0,1],

            # Perlengkapan jaringan
            "Line Post;24kV;12.5kN;Polymer SIR": [0,1,2,3,4,5,6],
            "INSULATED TOP TIES": [0,1,2,3,4,5,6],

            # Pelindung/pengontrol
            "LA;24-24kV;1kA;Polymer": [0,1,2,3],
            "CUT OUT;24kV;6-100A;POLYMER-125kV": [0,1,2,3],
            "CUT OUT ACC;FUSE LINK 26 kV 8A": [19980]
        }
    }
}


# Urutan layer agar tampilan tidak tumpang tindih salah
urutan_layer_per_gardu = {
    "Gardu Cantol": [
        "Pembuatan pondasi gardu Cantol",
        "Pasang Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat",
        "TIANG BETON BULAT 12m/350 daN (terpasang) + Lansir",
        "PENGADAAN DAN PEMASANGAN PENGHALANG PANJAT",
        "Pasang rangka gardu Cantol",
        "Pasang TM - 1 B/H ( TIANG TUMPU )",
        "Pasang Pentanahan TM Gardu Cantol",
        "Pemasangan hardware, isolator, pipa, PGTM/TR Cantol",
        "LVSB;DIST;3P;400V;250A;2Line;OH",
        "Pasang trafo 100 kVA",
        "Pasang Arrester + Jumper",
        "Pasang Cut Out + Jumper",
        "Pasang Cover Bushing Trafo",
        "Pembuatan Nama dan logo gardu",
        "Penomoran Tiang TM",
        "PENGADAAN DAN PEMASANGAN TANDA KILAT KECIL"
    ],
    "Gardu Tembok": [
        "PLAT LANTAI KABEL LKP PENUTUP LUBANG DENGAN AQUAPROOF",
        "PEMASANGAN CUBICLE PROTEKSI TRAFO (PB)",
        "LVSB;DIST;3P;380V;1000A;6LINE;ID",
        "PASANG PHBTR INDOOR 4 JURUSAN / 6 JURUSAN",
        "PASANG TRAFO 400 kVA GARDU TEMBOK",
        "PEMASANGAN KABEL NYY DARI TRAFO KE PHBTR DI GARDU TEMBOK",
        "PEMASANGAN INSTALASI PENTANAHAN DI GARDU TEMBOK"

    ],
    "Gardu Portal": [
        "Pembuatan pondasi gardu Portal",
        "Pasang rangka gardu Portal",
        "TIANG BETON BULAT 12m/350 daN (terpasang) + Lansir",
        "Pasang Kabel Naik TR & Pipa Pelindung untuk Jurusan Baru lengkap perbaikan Rabat",
        "Pasang TM - 1 B/H ( TIANG TUMPU )",
        "Pemasangan hardware, isolator, pipa, PGTM/TR Portal",
        "LVSB;DIST;3P;400V;630A;4Line;OD",
        "Pasang trafo 200 - 250 kVA",
        "Pasang Arrester + Jumper",
        "Pasang Cut Out + Jumper",
        "Pasang Cover Arrester",
        "Pasang Cover Cut Out",
        "Pasang Cover Bushing Trafo",
        "Pasang Pentanahan Ganda TM Gardu Portal",
        "Pembuatan Nama dan logo gardu",
        "Penomoran Tiang TM",
        "PENGADAAN DAN PEMASANGAN PENGHALANG PANJAT",
        "PENGADAAN DAN PEMASANGAN TANDA KILAT KECIL"
    ]
}
# =========================
# INISIALISASI & TOMBOL RESET
# =========================

if "reset_trigger" not in st.session_state:
    st.session_state.reset_trigger = False

def reset_all():
    st.session_state.df_dropdown_state["KEBUTUHAN"] = 0
    st.session_state.df_numeric_state["KEBUTUHAN"] = 0
    
if st.button("Reset Kebutuhan"):
        st.session_state.reset_trigger = True
        st.rerun()
    
# =========================
# PILIH GARDU
# =========================
with st.sidebar:
    st.header("Pilih Tipe Gardu")
    selected_gardu = st.selectbox("Tipe Gardu", list(gardu_data.keys()))

# Ambil dict material (handle kasus "Gardu Garpor" yang pakai key berbeda)
material_key = "material" if "material" in gardu_data[selected_gardu] else "material dan jasa"
material_dict = gardu_data[selected_gardu][material_key]
image_file = gardu_data[selected_gardu]["image"]

# =========================
# LOGIKA PEMISAHAN INPUT
# =========================
def is_length_based(item_name: str) -> bool:

    if selected_gardu.lower() == "gardu tembok":
        return False
    keywords = ["CABLE", "CONDUCTOR", "N2XSY", "NYY", "AAAC"]
    upper_name = item_name.upper()
    return any(k in upper_name for k in keywords)

# Buat DataFrame awal
df_input = pd.DataFrame({
    "NAMA MATERIAL": list(material_dict.keys()),
    "KEBUTUHAN": [0] * len(material_dict)
})

# Bagi material berdasarkan jenis input
dropdown_items = [m for m in df_input["NAMA MATERIAL"] if not is_length_based(m)]
numeric_items  = [m for m in df_input["NAMA MATERIAL"] if is_length_based(m)]


df_dropdown = pd.DataFrame({
    "NAMA MATERIAL": dropdown_items, 
    "KEBUTUHAN": [0.0] * len(dropdown_items)
})
df_numeric  = pd.DataFrame({
    "NAMA MATERIAL": numeric_items,  
    "KEBUTUHAN": [0.0] * len(numeric_items)
})


# Inisialisasi state untuk menyimpan input
if "df_dropdown_state" not in st.session_state:
    st.session_state.df_dropdown_state = df_dropdown.copy()

if "df_numeric_state" not in st.session_state:
    st.session_state.df_numeric_state = df_numeric.copy()

if "reset_trigger" not in st.session_state:
    st.session_state.reset_trigger = False

# Pilihan dropdown sesuai gardu
if selected_gardu == "Gardu Portal":
    pilihan_jumlah = [0, 1, 2, 3, 4, 5, 6]
elif selected_gardu == "Gardu Cantol":
    pilihan_jumlah = [0, 1, 2, 3]
else:
    pilihan_jumlah = [x * 0.5 for x in range(2, 201)]  # 1.0, 1.5, 2.0, ..., 100.0


# =========================
# INPUT USER
# =========================
col_input, col_vis = st.columns([1.2, 1])
with col_input:
    st.markdown(f"### Input Kebutuhan Material - {selected_gardu}")
    if st.session_state.reset_trigger:
            st.session_state.df_dropdown_state["KEBUTUHAN"] = 0
            st.session_state.df_numeric_state["KEBUTUHAN"] = 0
            st.session_state.reset_trigger = False
    if len(df_dropdown):
        st.markdown("**Jenis Pekerjaan / Material (Tabel Volume Pemborong)**")
        edited_dropdown = st.data_editor(
            df_dropdown,
            use_container_width=True,
            column_config={
                "KEBUTUHAN": st.column_config.SelectboxColumn(
                "KEBUTUHAN", options=pilihan_jumlah, default=0
                ),
                "NAMA MATERIAL": st.column_config.TextColumn("NAMA MATERIAL", disabled=True)
            },
            hide_index=True
        )
    else:
        edited_dropdown = pd.DataFrame(columns=df_dropdown.columns)
    if len(df_numeric):            
        st.markdown("Material Kabel / Konduktor (Tabel Volume PLN )")
        edited_numeric = st.data_editor(
                df_numeric,
                use_container_width=True,
                column_config={
                    "KEBUTUHAN": st.column_config.NumberColumn("KEBUTUHAN", min_value=0, step=1),
                    "NAMA MATERIAL": st.column_config.TextColumn("NAMA MATERIAL", disabled=True)
                },
            hide_index=True
        )
    else:
        edited_numeric = pd.DataFrame(columns=df_numeric.columns)

    # Gabungkan kembali & urut sesuai df_input awal
    edited_df = pd.concat([edited_dropdown, edited_numeric], ignore_index=True)
    edited_df = edited_df.set_index("NAMA MATERIAL").reindex(df_input["NAMA MATERIAL"]).reset_index()
    

# =========================
# VISUALISASI DINAMIS
# =========================
with col_vis:
    st.subheader("Visualisasi Komponen Sesuai Kebutuhan")

    gambar_material = gambat_material_per_gardu.get(selected_gardu, {})
    urutan_layer = urutan_layer_per_gardu.get(selected_gardu, list(gambar_material.keys()))

    canvas_size = (600, 600)
    canvas = Image.new("RGBA", canvas_size, (0, 0, 0, 0))
    if selected_gardu == "Gardu Tembok":    
        try:
            dasar_path = "Bagian_Gardu/GarduTembok/GARDU TEMBOK-Pondasi GT.drawio.png"  # ganti dengan path asli
            dasar_layer = Image.open(dasar_path).convert("RGBA").resize(canvas_size)
            canvas = Image.alpha_composite(canvas, dasar_layer)
        except Exception as e:
            st.warning(f"Gagal load gambar dasar gardu tembok: {e}")

    kebutuhan_terisi = pd.to_numeric(edited_df["KEBUTUHAN"], errors="coerce").fillna(0).astype(float).sum()

    if kebutuhan_terisi == 0:
        st.warning("Silakan masukkan jumlah kebutuhan pada tabel untuk melihat visualisasi komponen gardu.")
    else:
        for nama in urutan_layer:
            if nama in gambar_material:
                jumlah = int(edited_df.loc[edited_df["NAMA MATERIAL"] == nama, "KEBUTUHAN"].fillna(0).values[0])
                path_gambar = gambar_material[nama]

                if isinstance(path_gambar, str):
                    path_gambar = [path_gambar]
                gambar_terpakai = path_gambar[:jumlah]

                for path in gambar_terpakai:
                    try:
                        layer = Image.open(path).convert("RGBA").resize(canvas_size)
                        canvas = Image.alpha_composite(canvas, layer)
                    except Exception as e:
                        st.warning(f"Gagal load gambar: {nama} ({path}) → {e}")
            
    if canvas:
        buf = io.BytesIO()
        canvas.save(buf, format="PNG")
        st.image(buf.getvalue(), caption=f"Rekonstruksi {selected_gardu}", width=400)

import openpyxl
from io import BytesIO
from openpyxl.styles import Font, Alignment

# Load template Excel dari dictionary atau path lokal
template_path = "Bagian_Gardu/TemplateRAB_Percobaan1.xlsx"  # sesuaikan lokasi jika dari dictionary

# Hitung ulang total harga
edited_df["HARGA SATUAN"] = edited_df["NAMA MATERIAL"].map(material_dict)
#edited_df["KEBUTUHAN"] = pd.to_numeric(edited_df["KEBUTUHAN"], errors="coerce").fillna(0)
edited_df["KEBUTUHAN"] = pd.to_numeric(edited_df["KEBUTUHAN"], errors="coerce").fillna(0).astype(float)
edited_df["TOTAL HARGA"] = edited_df["KEBUTUHAN"] * edited_df["HARGA SATUAN"]

df = edited_df.copy()
total_anggaran = df["TOTAL HARGA"].sum()
# =============== #
# TAMBAHKAN KE XLSX
# =============== #
template_path = "Bagian_Gardu/TemplateFormatRAB.xlsx"
output = BytesIO()

# Load template
wb = load_workbook(template_path)
ws = wb.active

kategori_map = {
    "JASA SUTM": ["Pasang TM - 1 B/H ( TIANG TUMPU )"],
    "PEMASANGAN INSTALASI GARDU TEMBOK": ["PEMASANGAN CUBICLE PROTEKSI TRAFO (PB)"],
    "PEMASANGAN TRAFO DISTRIBUSI": ["PASANG TRAFO 400 kVA GARDU TEMBOK"],
    "PENGADAAN MATERIAL ASSESORIES SUTM & GARDU TIANG":["CONDUCTOR;CU;50MM2;19.28KN"],
    "JASA GARDU DISTRIBUSI": ["Pasang rangka gardu"],
    "LAIN - LAIN": ["Tagging GPS", "PEMBUATAN GAMBAR PELAKSANAAN PADA KERTAS HVS 80 MGs UKURAN A3 ( 4 LEMBAR )"],
    "TIANG BETON": ["TIANG BETON BULAT 12m/350 daN (terpasang) + Lansir"],
    "MATERIAL KABEL": ["CONDUCTOR;AAAC-S;"],
    "MATERIAL KONEKTOR": ["CABLE PWR ACC;CABLE SHOE CU ID 1H 70mm2"],
    "MATERIAL PERLENGKAPAN GAARDU": ["TRF DIS;D3;", "TRA DIS;24kV;"],
    "MATERIAL PERLENGKAPAN JARINGAN": ["Line Post;24kV;12,5kN"],
    "MATERIAL PELINDUNG/PENGONTROL": ["LA;20-24kV;K;10kA;POLYMER;;"],
    "MATERIAL PELINDUNG/PENGONTROL": ["FUSE;20/24kV;16A;TUBE;D24mm"]
}
kategori_map_kebutuhan = [
    "CABLE PWR;N2XSY;1X35mm2;20kV;UG",
    "TIANG BETON"
]

start_row = 11
current_row = start_row
current_category = None

def safe_float(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0

# Flag pindah kolom
pindah_ke_h = False  # letakkan sebelum loop for

for _, row in df.iterrows():
    nama_material = str(row["NAMA MATERIAL"]).lower()

    # Cek apakah material ini masuk kategori tertentu (substring match)
    for kategori, trigger_list in kategori_map.items():
        if any(trigger.lower() in nama_material for trigger in trigger_list) and kategori != current_category:
            # Tulis kategori pekerjaan
            ws[f'D{current_row}'] = kategori
            ws.merge_cells(f'D{current_row}:L{current_row}')
            ws[f'D{current_row}'].font = Font(bold=True)
            ws[f'D{current_row}'].alignment = Alignment(horizontal="left")
            ws[f'D{current_row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            current_row += 1
            current_category = kategori
            break

    # Tulis data material
    if any(trigger.lower() in nama_material for trigger in kategori_map_kebutuhan):
        pindah_ke_h = True  # aktifkan flag kalau ketemu patokan
        pindah_ke_k = True  # Patokan untuk total harga
        
    ws[f'D{current_row}'] = str(row["NAMA MATERIAL"])
    
    
    if pindah_ke_h:
        ws[f'H{current_row}'] = safe_float(row["KEBUTUHAN"])  # pindah kolom H mulai dari sini
        ws[f'K{current_row}'] = safe_float(row["TOTAL HARGA"])
    else:
        ws[f'I{current_row}'] = safe_float(row["KEBUTUHAN"])  # default kolom I
        ws[f'L{current_row}'] = safe_float(row["TOTAL HARGA"])
    
    #ws[f'I{current_row}'] = safe_float(row["KEBUTUHAN"])
    ws[f'J{current_row}'] = safe_float(row["HARGA SATUAN"])
    current_row += 1
# Tambahkan total anggaran ke KL60
try:
    ws["K60"] = total_anggaran
    ws["K60"].number_format = '#,##0'  # supaya format ribuan rapi
except ValueError:
    st.warning("Sel K60 tidak bisa diisi, pastikan format sel di Excel sesuai.")

# Tampilan DataFrame dan total anggaran
st.markdown(f"### **Total Anggaran: Rp {total_anggaran:,.0f}**")
st.markdown("### Tabel RAB")
st.dataframe(df, use_container_width=True)

# Simpan ke BytesIO
wb.save(output)
output.seek(0)  # <-- Penting!

# Tombol Download dengan styling
custom_button = f"""
    <style>
    .download-btn {{
        background: linear-gradient(90deg, #007bff, #00c6ff);
        color: white;
        padding: 12px 24px;
        border-radius: 12px;
        text-align: center;
        font-weight: bold;
        text-decoration: none;
        display: inline-block;
        transition: 0.3s;
    }}
    .download-btn:hover {{
        background: linear-gradient(90deg, #0056b3, #0096c7);
        transform: scale(1.05);
    }}
    </style>
"""

st.markdown(custom_button, unsafe_allow_html=True)

# Tombol Download
st.download_button(
    label="Download Laporan RAB (Excel)",
    data=output,
    file_name=f"laporan_rab_{selected_gardu.replace(' ', '_').lower()}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
